import time

from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.window import WindowTypes
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import selenium.common.exceptions
import psycopg2.errors
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import numpy as np
import openpyxl

# Путь к файлу Excel
filename = "news_cb_data.xlsx"

# Попытка открыть существующий файл или создание нового, если файл не существует
try:
    workbook = openpyxl.load_workbook(filename)
    print("Файл найден, добавляем данные.")
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    print("Файл не найден, создаем новый.")
    # Если создаем новый файл, можно сразу добавить заголовки столбцов
    sheet = workbook.active
    sheet['A1'] = 'id'
    sheet['B1'] = 'news_date'
    sheet['C1'] = 'news_text'
    sheet['D1'] = 'news_url'
    sheet['E1'] = 'news_information'
    workbook.save(filename)

sheet = workbook.active



class DBWORK:
    def __init__(self):
        self.url = None
        self.purchase_text = None
        self.connection = None
        self.cursor = None
        self.driver = None
        self.realisation_level = None
        self.sphere = None
        self.industry = None
        self.city = None
        self.dates = None
        self.purchase_id = None
        self.all_money = None
        self.purchase_region = None
        self.private_percent = None
        self.gov_percent = None



class PARSER(DBWORK):
    def __init__(self):
        super().__init__()

    def driver_connection(self):
        '''подключаем веб-драйвер'''
        options = Options()
        options.add_argument("--headless")
        '''options=options  можно вписать в ChromeDriverManager() чтобы браузер не открывался'''
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    def site_opening(self, url):
        return self.driver.get(url)

    def filling_out_forms(self, form, information):
        '''заполнение форм на сайте'''
        try:
            return form.send_keys(information)
        except:
            return

    def click(self, click_elem):
        try:
            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, click_elem))).click()
        except:
            return

    def searching(self, mode, elem):
        try:
            result = ''
            if mode == 'xpath':
                result = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.XPATH, elem)))
            elif mode == 'css':
                result = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, elem)))
            return result
        except:
            return


    def results_define(self, i, news_date, news_text, news_url, page_information):
        # print(news_date, news_text, news_url, page_information, sep='----------------------')\
        new_row = sheet.max_row + 1
        sheet.cell(row=new_row, column=1, value=i)
        sheet.cell(row=new_row, column=2, value=news_date)
        sheet.cell(row=new_row, column=3, value=news_text)
        sheet.cell(row=new_row, column=4, value=news_url)
        sheet.cell(row=new_row, column=5, value=page_information)

        # Сохранение файла после каждой итерации
        workbook.save(filename)
        for row in sheet.iter_rows(min_row=1, max_col=5, values_only=True):
            print(row)
    def back(self):
        self.driver.execute_script("window.history.go(-1)")

    def quit(self):
        self.driver.quit()
        self.cursor.close()
        self.connection.close()
    def close_window(self):
        self.driver.close()

    def script(self):
        self.driver.switch_to.new_window(WindowTypes.TAB)
    def c(self):
        self.driver.switch_to.window(self.driver.window_handles[0])

def main():
    pars_page = PARSER()
    pars_page.driver_connection()
    pars_page.site_opening("https://www.cbr.ru/news/")
    i = 2
    while True:
        try:
            pars_page.c()
            news_date = pars_page.searching('css', f'#events_tab100 > div.news-speeches_wrap.items_data > div:nth-child({i}) > div > div > div > div.news_date').text
            link = pars_page.searching('css', f'#events_tab100 > div.news-speeches_wrap.items_data > div:nth-child({i}) > div > div > a')
            news_url = link.get_attribute('href')
            news_text = link.text
            print(news_url)
            pars_page.script()
            pars_page.site_opening(news_url)
            i += 1
            try:
                page_information = pars_page.searching('css', '#content > div > div > div > div > div.landing-text')
            except:
                continue
            pars_page.results_define(i, news_date, news_text, news_url, page_information.text)
            pars_page.close_window()
        except:
            pars_page.click('#_buttonLoadNextEvt')
            continue



if __name__ == '__main__':
    main()
